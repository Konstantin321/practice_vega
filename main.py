from openpyxl import load_workbook

wb = load_workbook("vega.xlsx", data_only="NO")
ws = wb.active

rng_dicipline = ws['C2':chr(ord('A') + ws.max_column - 2) + '2'] # диапазон на дисциплины

dicipline_list = []

for row in rng_dicipline:
    dicipline_list.append(list(row))

dicipline_list = dicipline_list[0]

# вытаскиваю название дисциплины и ее протяженность
for i in range(len(dicipline_list)):
    dicipline_list[i] = list(dicipline_list[i].value.split(','))
    if '/' in dicipline_list[i][0]:
        dicipline_list[i] = [dicipline_list[i][0][:dicipline_list[i][0].find('/') - 2], list(dicipline_list[i][0].split())[-1], dicipline_list[i][1]]
    else:
        dicipline_list[i] = [dicipline_list[i][0], 'NULL', dicipline_list[i][1]]

#получаю имена
rng_name = ws['B3':'B' + str(ws.max_row)] # диапазон на фио
name_list = []

for i in range(len(rng_name)):
    name_list.append(list(rng_name[i])[0])

print(name_list[0])